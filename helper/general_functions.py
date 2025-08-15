import os
import string
import ast
import re
import csv
import ast
import numpy as np
import pandas as pd
from nltk import word_tokenize
from nltk.corpus import stopwords
import openpyxl


def split_text(text, max_length=300):
    text = clean_text(text)
    sentences = text.split('. ')
    chunks = []
    current_chunk = []
    try:
        for sentence in sentences:
            if len(' '.join(current_chunk + [sentence])) > max_length:
                chunks.append(' '.join(current_chunk))
                current_chunk = [sentence]
            else:
                current_chunk.append(sentence)
        if current_chunk:
            chunks.append(' '.join(current_chunk))
    except:
        print("split_text", text)
    return chunks

def create_and_write_csv(file_name, data):   
    dictory_path = "feature"
    filename = os.path.join(dictory_path, file_name + '.csv')
    with open(filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Key', 'Array'])
        for key, value in data.items():
            if isinstance(value, np.ndarray):
                array_str = str(value.tolist())
            else:
                array_str = str(value)
            writer.writerow([key, array_str])

def load_data_from_csv(file_path):
    data = {}
    with open(file_path, 'r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row in csv_reader:
            key = row['Key']
            array_str = row['Array']
            if array_str.startswith('[') and array_str.endswith(']'):
                data[key] = np.array(eval(array_str))
            else:
                array = np.array(ast.literal_eval(array_str))
                data[key] = array
    return data

def parse_array_from_string(array_string):
    try:
        if isinstance(array_string, (int, float)):
            return [float(array_string)]

        array_string = array_string.strip()
        array_string = re.sub(r'(?<![\d.])e[\d.]+', '', array_string)
        return ast.literal_eval(array_string)
    except (ValueError, SyntaxError):
        return []
    
def read_csv_file(csv_file):
    keys = []
    values = []

    with open(csv_file, 'r', encoding='utf-8') as file:
        csv_reader = csv.reader(file)
        next(csv_reader)  # Bỏ qua header
        for row in csv_reader:
            key = row[0]  # ID là cột đầu tiên
            value_str = row[1]  # Giá trị là cột thứ hai
            value = ast.literal_eval(value_str)  # Chuyển đổi chuỗi thành vector
            keys.append(key)
            values.append(np.array(value))

    return keys, values


def softmax(x):
    """Compute the softmax of vector x.
    """
    exp_x = np.exp(x)
    softmax_x = exp_x / np.sum(exp_x)
    return softmax_x


def sigmoid(x):
    x = np.clip(x, -709, 709)
    s = 1 / (1 + np.exp(-x))
    return s

stop_words = stopwords.words("english") + list(string.punctuation)
def word_segment(text):
    if pd.isnull(text):
        return [] 
    word_seg = word_tokenize(str(text).lower()) 
    
    return word_seg

def preprocessed(text):
    return text.split("\.")

def clean_text(text):
    text = re.sub(r'\.{2,}', ' ', text)
    return text
    
def format_array(arr):
    return "[" + ", ".join(map(str, arr)) + "]"

def convert_string_to_float_list(string):
    try:
        return np.array(ast.literal_eval(string), dtype=np.float64)
    except:
        return np.array([])

def save_to_excel(values, headers, output_path):
    if os.path.exists(output_path):
        workbook = openpyxl.load_workbook(output_path)
        sheet = workbook.active
        start_row = sheet.max_row
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        start_row = 1 

        for index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=index, value=header)
            
    for row_index, row_values in enumerate(values, start=start_row + 1):
        for col_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_index, column=col_index, value=round(value, 4))

    workbook.save(output_path)
    print(f"Đã lưu danh sách giá trị vào tệp '{output_path}' thành công.")